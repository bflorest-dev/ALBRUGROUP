import csv
import io
import json
import re
import subprocess
from collections import Counter
from datetime import datetime, date, time, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
DOCS = ROOT / "DOCUMENTORAIZ"
AUDITORIA = ROOT / "AUDITORIA"
OUT = ROOT / "PREBACKFILL"
OUT.mkdir(exist_ok=True)

POSTVENTA_XLSX = DOCS / "POSTVENTA_CLARO.xlsx"
VENTAS_XLSX = DOCS / "VENTAS_CLARO.xlsx"
PENDIENTES_CSV = AUDITORIA / "faltantes_excel.csv"
OUTPUT_JSON = OUT / "prebackfill_data.json"

ADMIN = {"usuario_id": 1, "empleado_id": 1, "nombre": "ADMINISTRADOR"}
CLARO_TEAM_ID = 2
CLARO_PROVEEDOR_ID = 2


def normalize(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return re.sub(r"\s+", "", text)


def clean_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def excel_date(value):
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        # Excel serial date, Windows epoch.
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date().isoformat()
    text = clean_text(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text


def excel_datetime(value):
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return datetime.combine(value, time.min).isoformat(sep=" ", timespec="seconds")
    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).isoformat(sep=" ", timespec="seconds")
    text = clean_text(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.isoformat(sep=" ", timespec="seconds")
        except ValueError:
            pass
    return text


def excel_money(value):
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float, Decimal)):
        return f"{Decimal(str(value)).quantize(Decimal('0.01'))}"
    text = str(value).strip()
    text = text.replace("S/.", "").replace("S/", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    text = re.sub(r"[^0-9.\-]", "", text)
    if not text:
        return ""
    try:
        return f"{Decimal(text).quantize(Decimal('0.01'))}"
    except InvalidOperation:
        return str(value).strip()


def split_coords(value):
    text = clean_text(value)
    if not text:
        return "", ""
    parts = [p.strip() for p in text.split(",")]
    if len(parts) >= 2:
        return parts[0], parts[1]
    return "", ""


def read_xlsx_table(path, sheet_names, header_row):
    wb = load_workbook(path, data_only=True, read_only=True)
    rows = []
    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = {}
        seen_headers = {}
        for cell in ws[header_row]:
            if cell.value is not None:
                header = clean_text(cell.value)
                seen_headers[header] = seen_headers.get(header, 0) + 1
                if seen_headers[header] > 1:
                    header = f"{header}__{seen_headers[header]}"
                headers[cell.column] = header
        for row in ws.iter_rows(min_row=header_row + 1):
            item = {"_sheet": sheet_name, "_row": row[0].row}
            non_empty = False
            for cell in row:
                header = headers.get(cell.column)
                if not header:
                    continue
                item[header] = cell.value
                if cell.value not in (None, ""):
                    non_empty = True
            if non_empty:
                rows.append(item)
    return rows


def read_csv(path):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def docker_csv(db_container, db_name, sql):
    copy_sql = f"COPY ({sql.rstrip().rstrip(';')}) TO STDOUT WITH CSV HEADER"
    cmd = [
        "docker",
        "exec",
        "-i",
        db_container,
        "psql",
        "-U",
        "postgres",
        "-d",
        db_name,
        "-c",
        copy_sql,
    ]
    result = subprocess.run(cmd, text=True, capture_output=True)
    if result.returncode != 0:
        raise RuntimeError(result.stderr or result.stdout)
    return list(csv.DictReader(io.StringIO(result.stdout)))


def infer_plan_family(servicio):
    text = clean_text(servicio).upper().replace(" ", "")
    if "3PLAY" in text:
        return "3 PLAY INTER + TLF + TV"
    if "TV" in text:
        return "2 PLAY INTERNET + TV"
    if "TELF" in text or "TLF" in text:
        return "2 PLAY INTER + TLF"
    if "1PLAY" in text:
        return "1 PLAY INTERNET"
    return ""


def choose_plan(servicio, precio, plans):
    price = excel_money(precio)
    family = infer_plan_family(servicio)
    same_price = [p for p in plans if excel_money(p["precio"]) == price]
    same_family = [p for p in same_price if p["nombre"] == family]
    active_same_family = [p for p in same_family if clean_text(p["activo"]).lower() in ("t", "true", "1")]
    active_same_price = [p for p in same_price if clean_text(p["activo"]).lower() in ("t", "true", "1")]
    chosen = None
    confidence = "SIN_MATCH"
    note = ""
    if active_same_family:
        chosen = active_same_family[0]
        confidence = "PRECIO_Y_FAMILIA_ACTIVO"
    elif same_family:
        chosen = same_family[0]
        confidence = "PRECIO_Y_FAMILIA_INACTIVO"
    elif active_same_price:
        chosen = active_same_price[0]
        confidence = "PRECIO_ACTIVO"
        note = "Sin familia exacta; elegido por cargo fijo CLARO activo."
    elif same_price:
        chosen = same_price[0]
        confidence = "PRECIO_INACTIVO"
        note = "Sin familia exacta; elegido por cargo fijo CLARO inactivo."
    if chosen is None:
        return {
            "plan_id": "",
            "plan_nombre": "",
            "plan_precio": price,
            "plan_confianza": confidence,
            "plan_alerta": "No existe plan CLARO con ese cargo fijo.",
        }
    if len(same_price) > 1 and not note:
        note = f"Cargo fijo compartido por {len(same_price)} planes CLARO; se desempato por familia."
    return {
        "plan_id": chosen["id"],
        "plan_nombre": chosen["nombre"],
        "plan_precio": excel_money(chosen["precio"]),
        "plan_confianza": confidence,
        "plan_alerta": note,
    }


def payment_decision(source, amount, paid_at):
    marker = clean_text(source).upper()
    amount = excel_money(amount)
    paid_at = excel_date(paid_at)
    if not marker:
        return {
            "periodo_estado": "ABIERTO",
            "pago_aportante": "",
            "pago_estado": "",
            "pago_monto": "",
            "fecha_pago": "",
            "cliente_estado": "",
        }
    if "BAJA" in marker:
        return {
            "periodo_estado": "CERRADO_BAJA",
            "pago_aportante": "",
            "pago_estado": "",
            "pago_monto": "",
            "fecha_pago": "",
            "cliente_estado": "BAJA",
        }
    if "ALBRU" in marker or "EMPRESA" in marker:
        return {
            "periodo_estado": "CERRADO_PAGO_EMPRESA",
            "pago_aportante": "EMPRESA",
            "pago_estado": "PAGADO_EMPRESA",
            "pago_monto": amount,
            "fecha_pago": paid_at,
            "cliente_estado": "ACTIVO",
        }
    if "CLIENTE" in marker:
        return {
            "periodo_estado": "CERRADO_PAGO_CLIENTE",
            "pago_aportante": "CLIENTE",
            "pago_estado": "PAGADO_CLIENTE",
            "pago_monto": amount,
            "fecha_pago": paid_at,
            "cliente_estado": "ACTIVO",
        }
    return {
        "periodo_estado": "REVISAR",
        "pago_aportante": "",
        "pago_estado": "",
        "pago_monto": amount,
        "fecha_pago": paid_at,
        "cliente_estado": "",
    }


def main():
    pendientes = read_csv(PENDIENTES_CSV)
    ventas = read_xlsx_table(VENTAS_XLSX, ["_PREVENTAS_", "PREVENTAS"], 1)
    postventa = read_xlsx_table(POSTVENTA_XLSX, ["JUN26", "JUL26", "AGO26"], 5)
    plans = read_csv(OUT / "planes_claro.csv")
    current = read_csv(OUT / "leads_actuales.csv")
    current_by_keys = {}
    for row in current:
        for key in (normalize(row.get("lead")), normalize(row.get("sot")), normalize(row.get("numero_documento_titular_servicio_snapshot"))):
            if key:
                current_by_keys.setdefault(key, row)

    ventas_by_key = {}
    for row in ventas:
        key = (normalize(row.get("LEAD")), normalize(row.get("SOT")), normalize(row.get("N° DOCUMENTO")))
        ventas_by_key[key] = row

    postventa_by_key = {}
    for row in postventa:
        key = (normalize(row.get("LEAD")), normalize(row.get("SOT")), normalize(row.get("N° DOCUMENTO")))
        postventa_by_key[key] = row

    details = []
    alerts = []
    for p in pendientes:
        key = (normalize(p.get("lead_excel")), normalize(p.get("sot_excel")), normalize(p.get("documento_excel")))
        v = ventas_by_key.get(key, {})
        pv = postventa_by_key.get(key, {})
        existing = current_by_keys.get(key[0]) or current_by_keys.get(key[1]) or current_by_keys.get(key[2]) or {}
        exists = bool(existing.get("id_lead"))
        lat, lon = split_coords(v.get("COORDENADAS"))
        fecha_inst = excel_date(pv.get("FECHA INSTALACIÓN") or p.get("fecha_instalacion_excel"))
        registro_at = excel_datetime(v.get("Marca temporal"))
        fecha_prog = excel_date(v.get("PROGRAMACION BO") or v.get("FECHA PROGRAMACION"))
        plan = choose_plan(v.get("SERVICIO A CONTRATAR"), v.get("PRECIO REGULAR"), plans)

        p1 = payment_decision(pv.get("PAGO R1"), pv.get("MONTO"), pv.get("FEC PAGO"))
        p2 = payment_decision(pv.get("PAGO R2"), pv.get("MONTO__2"), pv.get("FEC PAGO__2"))

        cliente_estado = "BAJA" if p1["cliente_estado"] == "BAJA" or p2["cliente_estado"] == "BAJA" else clean_text(pv.get("STATUS CLIENTE") or "ACTIVO").upper()
        final_estado = "NUEVO"
        action = "ACTUALIZAR_EXISTENTE" if exists else "CREAR_DESDE_CERO"

        row_alerts = []
        if not fecha_inst:
            row_alerts.append("Sin fecha instalacion postventa")
        if not lat or not lon:
            row_alerts.append("Coordenadas no separables")
        if plan["plan_confianza"] in ("SIN_MATCH", "PRECIO_ACTIVO", "PRECIO_INACTIVO"):
            row_alerts.append(plan["plan_confianza"])
        ubigeo = normalize(v.get("UBIGEO"))
        if len(ubigeo) not in (2, 4, 6):
            row_alerts.append("Ubigeo nacimiento revisar")
        if p1["periodo_estado"] == "REVISAR" or p2["periodo_estado"] == "REVISAR":
            row_alerts.append("Pago con marcador no reconocido")

        detail = {
            "hoja_postventa": p.get("hoja"),
            "fila_postventa": p.get("row_excel"),
            "estado_revision_previa": p.get("estado_revision"),
            "accion_backfill": action,
            "id_lead_existente": existing.get("id_lead", ""),
            "lead": key[0],
            "sec": normalize(pv.get("SEC") or p.get("sec_excel") or v.get("SEC")),
            "sot": key[1],
            "documento": key[2],
            "tipo_documento": clean_text(pv.get("TIPO DOCUMENTO") or v.get("TIPO DOCUMENTO")),
            "nombre_cliente": clean_text(pv.get("NOMBRES Y APELLIDOS CLIENTE") or v.get("NOMBRE COMPLETO CLIENTE")),
            "telefono_referencia": normalize(v.get("TELF REFERENCIA")),
            "correo": clean_text(v.get("CORREO ELECTRONICO")),
            "ubigeo_nacimiento_raw": ubigeo,
            "ubigeo_nacimiento_regla": "USAR_DIRECTO" if len(ubigeo) == 6 else "INFERIR_COMPLETANDO",
            "direccion": clean_text(v.get("DIRECCION INSTALACION")),
            "referencia": clean_text(v.get("REFERENCIAS")),
            "latitud": lat,
            "longitud": lon,
            "plano": clean_text(v.get("PLANO")),
            "piso": clean_text(v.get("PISO")),
            "interior": clean_text(v.get("INTERIOR")),
            "fecha_eventos_preventa": registro_at,
            "fecha_eventos_venta": registro_at,
            "fecha_programacion": fecha_prog,
            "fecha_instalacion": fecha_inst,
            "equipo_final": existing.get("id_equipo") or str(CLARO_TEAM_ID),
            "campana_final": existing.get("id_campana") or "",
            "base_final_si_nuevo": "" if exists else "REFERIDO",
            "actor_id": str(ADMIN["empleado_id"]),
            "actor_nombre": ADMIN["nombre"],
            "lead_etapa_final": "POSTVENTA",
            "lead_estado_final": final_estado,
            "cliente_estado_final": cliente_estado,
            "servicio_excel": clean_text(v.get("SERVICIO A CONTRATAR")),
            "precio_regular_excel": excel_money(v.get("PRECIO REGULAR")),
            "plan_id_resuelto": plan["plan_id"],
            "plan_nombre_resuelto": plan["plan_nombre"],
            "plan_precio_resuelto": plan["plan_precio"],
            "plan_confianza": plan["plan_confianza"],
            "plan_alerta": plan["plan_alerta"],
            "p1_emision": excel_date(pv.get("FEC. EMISION")),
            "p1_vencimiento": excel_date(pv.get("FEC. VENC")),
            "p1_monto": excel_money(pv.get("MONTO")),
            "p1_marcador_pago": clean_text(pv.get("PAGO R1")),
            "p1_estado_periodo": p1["periodo_estado"],
            "p1_aportante": p1["pago_aportante"],
            "p1_estado_pago": p1["pago_estado"],
            "p1_fecha_pago": p1["fecha_pago"],
            "p2_emision": excel_date(pv.get("FEC. EMISION__2")),
            "p2_vencimiento": excel_date(pv.get("FEC. VENC__2")),
            "p2_monto": excel_money(pv.get("MONTO__2")),
            "p2_marcador_pago": clean_text(pv.get("PAGO R2")),
            "p2_estado_periodo": p2["periodo_estado"],
            "p2_aportante": p2["pago_aportante"],
            "p2_estado_pago": p2["pago_estado"],
            "p2_fecha_pago": p2["fecha_pago"],
            "alertas": "; ".join(row_alerts),
        }
        details.append(detail)
        for alert in row_alerts:
            alerts.append({"lead": key[0], "sot": key[1], "alerta": alert})

    summary = {
        "total_pendientes": len(details),
        "crear_desde_cero": sum(1 for d in details if d["accion_backfill"] == "CREAR_DESDE_CERO"),
        "actualizar_existente": sum(1 for d in details if d["accion_backfill"] == "ACTUALIZAR_EXISTENTE"),
        "por_corte": Counter(f"{d['hoja_postventa']}|{d['estado_revision_previa']}" for d in details),
        "por_plan_confianza": Counter(d["plan_confianza"] for d in details),
        "alertas": Counter(a["alerta"] for a in alerts),
    }

    mapping = [
        ["Contacto", "lead", "POSTVENTA_CLARO.xlsx", "C", "Fuente principal"],
        ["Lead", "lead/sec/sot/documento", "POSTVENTA_CLARO.xlsx", "C/D/E/G", "Fuente principal"],
        ["Lead", "etapa/estado", "Regla", "POSTVENTA/NUEVO", "Estado NUEVO por decision operativa"],
        ["Lead", "asesor", "auth_db", "ADMIN", "usuario=1 empleado=1 ADMINISTRADOR"],
        ["Lead", "equipo", "Regla/BD", "existente o ClaroTeam=2", "Si existe se conserva"],
        ["Lead", "campana/base", "Regla/BD", "existente o sin campana + REFERIDO", "Solo para nuevos"],
        ["DatosPreventa", "tipo/doc/nombre", "VENTAS_CLARO.xlsx + POSTVENTA", "K/L/M + F/G/H", "Postventa valida identidad"],
        ["DatosPreventa", "ubigeo nacimiento", "VENTAS_CLARO.xlsx", "P", "Inferir si tiene 2/4 digitos"],
        ["Direccion", "direccion/coordenadas", "VENTAS_CLARO.xlsx", "AW/AZ/BA/BB/AX/AY", "Separar lat,long"],
        ["Plan", "plan/precio", "VENTAS_CLARO.xlsx", "BL+BN", "Prioridad cargo fijo regular CLARO"],
        ["Eventos PREVENTA", "REGISTRO/ASIGNACION/TIPIFICACION", "VENTAS_CLARO.xlsx", "A", "PREVENTA/COMPLETA"],
        ["Eventos VENTA", "INGRESADO/PROGRAMADO/INSTALADO", "VENTAS_CLARO.xlsx + POSTVENTA", "A/E/K", "Instalacion desde postventa K"],
        ["Postventa P1", "factura/pago/baja", "POSTVENTA_CLARO.xlsx", "BE/BF/BD/BG/BC", "BC define CLIENTE/EMPRESA/BAJA"],
        ["Postventa P2", "factura/pago/baja", "POSTVENTA_CLARO.xlsx", "BJ/BK/BI/BL/BH", "BH define CLIENTE/EMPRESA/BAJA"],
    ]

    data = {"summary": summary, "details": details, "alerts": alerts, "mapping": mapping}
    OUTPUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(OUTPUT_JSON)


if __name__ == "__main__":
    main()
